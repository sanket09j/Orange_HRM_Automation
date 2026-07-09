from openpyxl import load_workbook


class ExcelUtils:

    @staticmethod
    def get_row_count(file, sheet):

        workbook = load_workbook(file)

        worksheet = workbook[sheet]

        return worksheet.max_row

    @staticmethod
    def get_column_count(file, sheet):

        workbook = load_workbook(file)

        worksheet = workbook[sheet]

        return worksheet.max_column

    @staticmethod
    def read_data(file, sheet, row, column):

        workbook = load_workbook(file)

        worksheet = workbook[sheet]

        return worksheet.cell(row=row, column=column).value

    @staticmethod
    def write_data(file, sheet, row, column, data):

        workbook = load_workbook(file)

        worksheet = workbook[sheet]

        worksheet.cell(row=row, column=column).value = data

        workbook.save(file)